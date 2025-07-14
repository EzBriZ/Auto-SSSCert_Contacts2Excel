# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""

#Tie into SSSCert_Lineplan_Automation_lxml_etree
#this script aims to strip coordinates from SSS point contacts
#via shapefiles. 
#Takes coordinates and inputs them into SSS excel sheet

import geopandas as gpd
from openpyxl import load_workbook
import utm
import sqlite3

#import for button making
import tkinter as tk
from tkinter import filedialog, Label, simpledialog, messagebox

#initialize variable to test for button1 (browse) being executed before button2 (spacing)
button1_executed = False
button2_executed = False

#initialize start_column
start_row = 9

#function to allow user to find kml file they want to convert
def browse_shp():
    
    #initialize global variables to be used across functions
    global file_path
    global total_latlon
    global button1_executed
    
    #prompt user to select SSS contact shp file
    file_path = filedialog.askopenfilename(title="Select SHP File", 
                                           filetypes=[("SHP Files", "*.shp")]
                                           )
    #if they select a shp, open file and...
    if file_path:
        try:
            #read shapefile
            gdf = gpd.read_file(f"{file_path}")
            
            #initiate variable for later use
            total_latlon = []
            
            engrained_crs = gdf.crs
            print(engrained_crs)
     
            #variable to prompt user and record input of UTM Zone
            zone = simpledialog.askinteger("Input", "Enter UTM Zone:")
            

            #Look through metadata of shp and extract lat lon and line id
            for idx, row in gdf.iterrows():
                x = row.geometry.x
                y = row.geometry.y
                line_id = row['lineId']
                
                #convert UTM to lat lon
                lat, lon = utm.to_latlon(x, y, zone, northern=True)
                
                #Create them in a sorted list
                total_latlon.append((lat, lon, line_id))
          
            #read out that process has successfully completed
            label.config(text="Shapefile Successfully Found")
            
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}") 

    #store whether this button has been executed to test for right order of operations when calling button2
    button1_executed = True
    
    
def browse_hips():
    
    global headings_id
    global button2_executed
    
    if not button1_executed:
        #label to show file has been changed
        label.config(text="Error, Ensure you have browsed to shp file")
        return
    
    #Prompt for path to hips file
    file_path_sql = filedialog.askopenfilename(title="Select HIPS File",
                                                filetypes=[("HIPS SQLite Database", "*.hips")]
                                                )
    #Open up hips file and extract id and heading
    if file_path_sql:
        try:
            con = sqlite3.connect(file_path_sql)
            cursor = con.cursor()
            
            cursor.execute("SELECT concreteObjectId, heading FROM Line")

            headings_id = cursor.fetchall()
            con.close()
        
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}") 
    
    button2_executed = True
        
#Controls button to prompt user to find SSSCert excel sheet and where to save new copy
def browse_xlsx():
    
    #throw error if trying to execute button 2 without first executing button 1
    if not button1_executed:
        #label to show file has been changed
        label.config(text="Error, Ensure you have browsed to shp file")
        return
    
    if not button2_executed:
        #label to show file has been changed
        label.config(text="Error, Ensure you have browsed to shp file")
        return
    
    #prompt for filepath
    file_path_xlsx = filedialog.askopenfilename(title="Select SSSCert .xlsx File",
                                                filetypes=[("XLSX Files", "*.xlsx")]
                                                )
    
        
    #once given the SSSCert sheet
    if file_path_xlsx:
        try:
            #Load excel workbook
            workbook = load_workbook(filename=file_path_xlsx)

            #pick sheet
            sheet = workbook.active
            
            #create searchable dictionary to compare line id between the two lists - in order to pair up lines and insert correct headings
            dictionary = {line_id: heading for line_id, heading in headings_id}
            
            #Loop through and plug lat long into respective cells
            for i, (lat, lon, line_id) in enumerate(total_latlon):
                row = start_row + i
                #if more than 12 points were selected, end with last as to not break excel sheet
                if row > 20:
                    label.config(text="Operation Cancelled - Too many contacts selected (>12)")
                    return
                sheet.cell(row=row, column=2).value = lat
                sheet.cell(row=row, column=3).value = lon
                
                #Call back on dictionary and plug in heading 
                heading =  dictionary.get(line_id, "N/A")
                sheet.cell(row=row, column=4).value = heading
                                                         
            #ask where to save new file
            file_name = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                        title="Select Save File",
                                                        filetypes=[("XLSX Files", "*.xlsx")]
                                                        )
            #save the file
            workbook.save(file_name)
            
            #Confirm successful file save
            label.config(text=f"File saved to to: {file_name}")
            
            
        except Exception as e:
            messagebox.showerror("Error", f"An error occured: {e}")
            
            
# create tkinter GUI
root = tk.Tk()
root.geometry("500x200")
root.title("SSS Certification Auto Excel Sheet")

#set purpose of button1 (browse for shp)
browse_button = tk.Button(root, text="Browse for .shp", command=browse_shp)
browse_button.pack(pady=10)

#set purpose of button1 (browse for hips)
browse_button = tk.Button(root, text="Browse for .hips", command=browse_hips)
browse_button.pack(pady=10)

#set purpose of button2 (browse for and save new an excel sheet)
browse_button = tk.Button(root, text="Browse/Save .xlsx", command=browse_xlsx)
browse_button.pack(pady=10)


#set style of label
label = Label(root, text="")
label.pack(pady=20)

root.mainloop()
    